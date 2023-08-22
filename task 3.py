#!/usr/bin/env python
# coding: utf-8

# # Write a Python script using the openpyxl library to read data from an Excel spreadsheet, manipulate the data, and write the results back to a new spreadsheet.

# In[49]:


import pandas as pd


# In[50]:


from openpyxl import Workbook


# In[51]:


wb = Workbook()


# In[52]:


ws = wb.active


# In[80]:


Fruit_data=[["Fruit","Quantity"],["Apple",15],["Mango",20],["Apple",30],["Grape",5],["Banana",25],["Mango",35],["Pineapple",15],["Watermelon",27],["Grape",8],["Orange",17],["Cherry",25],["Watermelon",14],["Orange",39],["Peach",20],["Banana",30]]


# In[81]:


for row in Fruit_data:
    ws.append(row)


# In[82]:


wb.save("C:\\Users\\pooja yadav\\Documents\\PHN internship\\fruit data(task 3)\\Fruit data.xlsx")


# In[75]:


from openpyxl.styles import Font


# In[76]:


ft = Font(bold=True,name="Calibri",size=11,italic=False,color="27AE60")


# In[77]:


for col in ws["a1:b1"]:
    for cell in col:
        cell.font=ft


# In[78]:


from openpyxl.styles import PatternFill


# In[79]:


fill = PatternFill(fill_type="solid",start_color="D5F5E3",end_color="D5F5E3")


# In[61]:


ws_new = wb_new["Sheet1"]


# In[62]:


for row in ws_new["A1:B16"]:
    for cell in row:
        print("Cell Value -> " + str(cell.value))


# In[63]:


from openpyxl import load_workbook


# In[64]:


wb_new=load_workbook("C:\\Users\\pooja yadav\\Documents\\PHN internship\\total quantity of fruits(task 3)\\Total quantity of fruits.xlsx")


# In[65]:


ws_new = wb_new["Sheet1"]


# In[66]:


for row in ws_new["A1:B10"]:
    for cell in row:
        print("Cell Value -> " + str(cell.value))

